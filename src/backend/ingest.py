"""
ingest.py  —  Build the product index from the Excel catalog.

Run ONCE before starting the server:
    python ingest.py

Re-run whenever the Excel file changes OR after upgrading scikit-learn.
Creates: product_index.pkl
"""

import os
import pickle
import sklearn
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer

EXCEL_PATH = "ProductCatalog_tran.xlsx"
INDEX_PATH = "product_index.pkl"


def clean(value) -> str:
    return str(value).strip() if value is not None else ""


def build_document(row: tuple) -> str:
    return (
        f"Product: {clean(row[2])} "
        f"Type: {clean(row[3])} "
        f"Target diseases and pests: {clean(row[4])} "
        f"Applicable crops: {clean(row[6])} "
        f"Active ingredients: {clean(row[9])} "
        f"Usage and dosage: {clean(row[10])}"
    )


def build_metadata(row: tuple) -> dict:
    return {
        "product_id":   clean(row[0]),
        "name":         clean(row[2]),
        "product_type": clean(row[3]),
        "crops":        clean(row[6]),
        "spec":         clean(row[7]),
        "ingredients":  clean(row[9]),
        "usage":        clean(row[10]),
        "dilution":     clean(row[5]),
    }


def ingest():
    # ── Check if existing index was built with a different sklearn version ────
    if os.path.exists(INDEX_PATH):
        try:
            with open(INDEX_PATH, "rb") as f:
                old = pickle.load(f)
            saved_version = old.get("sklearn_version", "unknown")
            current_version = sklearn.__version__
            if saved_version != current_version:
                print(f"  sklearn version changed: {saved_version} → {current_version}")
                print(f"  Deleting old {INDEX_PATH} and rebuilding...")
                os.remove(INDEX_PATH)
            else:
                print(f"  {INDEX_PATH} is up to date (sklearn {current_version})")
                print("  Delete it manually and re-run if you want to force a rebuild.")
                return
        except Exception:
            print(f"  Could not read existing {INDEX_PATH} — rebuilding...")
            os.remove(INDEX_PATH)

    print(f"Loading: {EXCEL_PATH}")
    wb   = load_workbook(EXCEL_PATH, read_only=True)
    ws   = wb.active
    rows = [r for r in ws.iter_rows(values_only=True) if any(r)]
    data = rows[1:]
    print(f"  {len(data)} products found")

    docs  = [build_document(r) for r in data]
    metas = [build_metadata(r) for r in data]

    print("Building TF-IDF index...")
    vectorizer = TfidfVectorizer(max_features=2048, ngram_range=(1, 2), sublinear_tf=True)
    matrix     = vectorizer.fit_transform(docs)
    print(f"  Matrix: {matrix.shape}")

    with open(INDEX_PATH, "wb") as f:
        pickle.dump({
            "vectorizer":     vectorizer,
            "matrix":         matrix,
            "docs":           docs,
            "metas":          metas,
            "sklearn_version": sklearn.__version__,   # saved for version check
        }, f)

    print(f"\n✓ Saved to {INDEX_PATH}  (sklearn {sklearn.__version__})")


if __name__ == "__main__":
    ingest()