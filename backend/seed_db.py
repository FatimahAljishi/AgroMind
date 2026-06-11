"""
seed_db.py  —  Seed PostgreSQL with products, diseases, and mappings.

Run ONCE after the database is created:
    python seed_db.py

Safe to re-run — skips rows that already exist.

Seeds 3 tables:
  products         ← 114 rows from Excel catalog
  diseases         ← 20 rows  from disease_product_map.csv (unique disease names)
  disease_products ← 201 rows from disease_product_map.csv (the join table)
"""

import re
from numpy import random
import pandas as pd
from openpyxl import load_workbook
from database import SessionLocal, Product, Disease, DiseaseProduct, create_tables

EXCEL_PATH   = "ProductCatalog_tran.xlsx"
MAPPING_PATH = "disease_product_map.csv"


def clean(v) -> str:
    return str(v).strip() if v is not None else ""


def to_slug(name: str) -> str:
    """Convert disease name to a URL-safe ID, e.g. 'downy mildew' → 'downy-mildew'."""
    return re.sub(r"[^a-z0-9]+", "-", name.lower().strip()).strip("-")


def seed_products(db) -> dict:
    """Load Excel products. Returns {product_id: True} for existence checks."""
    wb   = load_workbook(EXCEL_PATH, read_only=True)
    ws   = wb.active
    rows = [r for r in ws.iter_rows(values_only=True) if any(r)][1:]

    added = skipped = 0
    existing = {p.product_id for p in db.query(Product.product_id).all()}

    for row in rows:
        pid = clean(row[0])
        if not pid or pid in existing:
            skipped += 1
            continue
        db.add(Product(
            product_id   = pid,
            name         = clean(row[2]),
            product_type = clean(row[3]),
            crops        = clean(row[6]),
            ingredients  = clean(row[9]),
            usage        = clean(row[10]),
            dilution     = clean(row[5]),
            spec         = clean(row[7]),
            price        = round(random.uniform(5, 250), 2),
        ))
        added += 1

    db.commit()
    print(f"  products  → {added} added, {skipped} skipped")
    return {p.product_id for p in db.query(Product.product_id).all()}


def seed_diseases_and_mappings(db, valid_product_ids: set):
    """Load disease names and 201 disease→product mappings from CSV."""
    df = pd.read_csv(MAPPING_PATH)

    # ── 1. Seed diseases table ────────────────────────────────────────────────
    unique_names   = df["disease_name"].str.strip().unique()
    existing_slugs = {d.id for d in db.query(Disease.id).all()}

    disease_added = 0
    for name in sorted(unique_names):
        slug = to_slug(name)
        if slug not in existing_slugs:
            db.add(Disease(id=slug, name=name.strip()))
            disease_added += 1

    db.commit()
    print(f"  diseases  → {disease_added} added, {len(unique_names)-disease_added} skipped")

    # Build slug lookup
    slug_map = {to_slug(name): name for name in unique_names}

    # ── 2. Seed disease_products join table ───────────────────────────────────
    existing_links = {
        (r.disease_id, r.product_id)
        for r in db.query(DiseaseProduct.disease_id, DiseaseProduct.product_id).all()
    }

    map_added = map_skipped = map_missing = 0
    for _, row in df.iterrows():
        disease_slug = to_slug(row["disease_name"])
        product_id   = str(row["product_id"]).strip()
        key          = (disease_slug, product_id)

        if key in existing_links:
            map_skipped += 1
            continue

        if product_id not in valid_product_ids:
            print(f"  WARNING: product {product_id} not in products table — skipping")
            map_missing += 1
            continue

        db.add(DiseaseProduct(disease_id=disease_slug, product_id=product_id))
        map_added += 1

    db.commit()
    print(f"  disease_products → {map_added} added, {map_skipped} skipped"
          + (f", {map_missing} missing product IDs" if map_missing else ""))


def seed():
    create_tables()
    db = SessionLocal()
    print("Seeding database...")
    valid_ids = seed_products(db)
    seed_diseases_and_mappings(db, valid_ids)
    db.close()
    print("\n✓ Seed complete")


if __name__ == "__main__":
    seed()
